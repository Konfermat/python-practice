import logging
import asyncio
from encodings import search_function

from multiprocessing import Queue
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pathlib import Path
import time

from openpyxl.workbook import Workbook

# библиотека для звука
from config import EXCEL_FILE, COLUMNS, STATUS_PENDING, STATUS_DONE, NOTIFY_FILE, CATEGORIES

logger = logging.getLogger(__name__)


class TaskManager:
    def __init__(self, file_event: asyncio.Event, analyzer_in_queue: Queue, analyzer_out_queue: Queue):
        self.tasks = []  # список asyncio.Task
        self.file_event = file_event  # Event at FileMonitor
        self.analyzer_in_queue = analyzer_in_queue
        self.analyzer_out_queue = analyzer_out_queue
        self.running = True

    def _play_sound(self):
        try:
            print('\a', end='')
        except Exception as e:
            logger.error(f'Ошибка при воспроизведении звука: {e}')

    async def _async_wait_and_notify(self, task_data: dict):
        task_time_str = task_data[COLUMNS['TIME']]
        task_desc = task_data[COLUMNS['TASK']]
        task_category = task_data[COLUMNS['CATEGORY']]
        row_index = task_data['row_index']
        try:
            target_time = datetime.strptime(task_time_str, '%H:%M').time()
            now = datetime.now()
            target_datetime = datetime.combine(now.date(), target_time)
            if target_datetime < now:
                target_datetime += timedelta(days=1)
            wait_seconds = (target_datetime - now).total_seconds()
            logger.info(
                f'Задача {task_desc} будет ждать {round(wait_seconds)} секунд до {target_time.isoformat(timespec='minutes')}.')
            await asyncio.sleep(wait_seconds)
            notify_msg = (
                f'НАПОМИНАНИЕ: [{task_time_str}] {task_category}. {task_desc}'
            )
            print('\n' + '-' * 30)
            print(notify_msg)
            print('-' * 30 + '\n')
            logger.info(f'Отправка уведомления о задаче  {notify_msg}')
            self._play_sound()
            logger.info('Воспроизведен звуковой сигнал')
            await asyncio.to_thread(self._update_status_excel, row_index, STATUS_DONE)
        except asyncio.CancelledError:
            logger.debug(f'Ассинхронное ожидание задачи {task_desc} отменено')
            raise  # перебрасываем ошибку для корректной обработки отмены
        except Exception as e:
            logger.error(f'Ошибка в асинхронном ожидании задачи {task_desc}: {e}')

    def _read_task_excel(self) -> list:
        tasks = []
        try:
            if not Path(EXCEL_FILE).exists():
                logger.warning(f'Файл {EXCEL_FILE} е найден. Создаю новый')
                self._create_excel_file()
                return []
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            required_columns = list(COLUMNS.values())
            for col in required_columns:
                if col not in headers:
                    logger.error(f'Нет обящательного поля {col}')
                    return tasks
            for i, row in enumerate(sheet.iter_rows(values_only=True, min_row=2), start=2):
                if not row or not row[0]:
                    continue
                task_dict = {
                    COLUMNS['TIME']: row[headers.index(COLUMNS['TIME'])],
                    COLUMNS['TASK']: row[headers.index(COLUMNS['TASK'])],
                    COLUMNS['CATEGORY']: row[headers.index(COLUMNS['CATEGORY'])] or ' не определена',
                    COLUMNS['STATUS']: row[headers.index(COLUMNS['STATUS'])] or STATUS_PENDING,
                    'row_index': i,
                }
                if task_dict[COLUMNS['STATUS']] == STATUS_PENDING:
                    tasks.append(task_dict)
                    logger.info(f'Задача загружена')
            workbook.close()
        except InvalidFileException as e:
            logger.error(f'Файл {EXCEL_FILE} поврежден')
        except Exception as e:
            logger.error(f'Ошибка при чтении {EXCEL_FILE}: {e}')

    def _create_excel_file(self):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Задачи'
            headers = list(COLUMNS.values())
            for i, h in enumerate(headers, start=1):
                sheet.cell(row=1, column=1, value=h)
            workbook.save(EXCEL_FILE)
            logger.info(f'Создан новый файл {EXCEL_FILE}')
        except Exception as e:
            logger.error(f'Ошибка при создании файла {EXCEL_FILE}: {e}')

    def _update_category_excel(self, row_index, new_category):
        try:
            with load_workbook(EXCEL_FILE) as wb:
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                category_col = headers.index(COLUMNS['CATEGORY']) + 1
                sheet.cell(row=row_index, column=category_col, value=new_category)
                wb.save(EXCEL_FILE)
                logger.info(f'Обновлена категория в строке {row_index} на {new_category}')
        except Exception as e:
            logger.error(f'Ошибка при обновлении категории: {e}')

    def _update_status_excel(self, row_index, new_status):
        try:
            with load_workbook(EXCEL_FILE) as wb:
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                status_col = headers.index(COLUMNS['STATUS']) + 1
                sheet.cell(row=row_index, column=status_col, value=new_status)
                wb.save(EXCEL_FILE)
                logger.info(f'Обновлен статус в строке {row_index} на {new_status}')
        except Exception as e:
            logger.error(f'Ошибка при обновлении статуса: {e}')

    def _add_task_console(self):
        try:
            print('Добавление новой задачи')
            while True:
                time_input = input('Введите время (ЧЧ:ММ): ').strip()
                try:
                    datetime.strftime(time_input, '%H:%M')
                    break
                except ValueError:
                    print('Неверный формат. Попробуй еще раз')
            task_desc = input('Введите задачу: ').strip()
            while not task_desc:
                print('Задача не может быть пустой')
                task_desc = input('Введите задачу: ').strip()
            self._add_task_excel(time_input, task_desc)
        except Exception as e:
            logger.error(f'Ошибка при добавлении задачи: {e}')

    def _add_task_excel(self, time_str, task_desc):
        try:
            if not Path(EXCEL_FILE).exists():
                self._create_excel_file()
            with load_workbook(EXCEL_FILE) as wb:
                sheet = wb.active
                max_row = sheet.max_row
                new_row = max_row + 1 if max_row > 1 else 2

                header = [cell.value for cell in sheet[1]]
                time_col = header.index(COLUMNS['TIME']) + 1
                time_col = header.index(COLUMNS['TASK']) + 1
                time_col = header.index(COLUMNS['CATEGORY']) + 1
                time_col = header.index(COLUMNS['STATUS']) + 1

                sheet.cell(row=new_row, column=time_col, value=time_str)
                sheet.cell(row=new_row, column=task_col, value=task_desc)
                sheet.cell(row=new_row, column=category_col, value='Не определена')
                sheet.cell(row=new_row, column=status_col, value=STATUS_PENDING)
                wb.save(EXCEL_FILE)
            logger.info(f'Добавлена новая задача в файл')
            self.analyzer_in_queue.put((new_row, task_desc))
        except Exception as e:
            logger.error(f'Ошибка при добавлении задачи в Excel: {e}')
            raise

    async def _process_analyzer_results(self):
        while self.running:
            try:
                if not self.analyzer_out_queue.empty():
                    row_index, new_category = self.analyzer_out_queue.get_nowait()
                    await asyncio.to_thread(self._update_category_excel, row_index, new_category)
                await asyncio.sleep(0.1)
            except Exception as e:
                logger.error(f'Ошибка при обработке результатов анализатора: {e}')

    async def _process_file_changes(self):
        while self.running:
            try:
                await self.file_event.wait()
                self.file_event.clear()
                logger.info("Обнаружено изменение файла. Перезагружаем задачи")

                # Отмена старых задач
                for task in self.tasks:
                    task.cancel()

                await self.load_and_schedule_tasks()

            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f'Ошибка при обработке изменений файла: {e}')

    async def load_schedule_tasks(self):
        tasks_data = await asyncio.to_thread(self._read_tasks_excel)
        for task in self.tasks:
            task.cancel()
        self.tasks.clear()
        for task_data in tasks_data:
            if task_data[COLUMNS['CATEGORY']] == 'Не определена':
                logger.info(f'Задача "{task_data[COLUMNS["TASK"]]}" требует анализа категории')
                self.analyzer_in_queue.put((task_data['row_index'], task_data[COLUMNS['TASK']]))

            task = asyncio.create_task(self._async_wait_and_notify(task_data))
            self.tasks.append(task)
            logger.info(f'Запланирована задача: {task_data[COLUMNS["TASK"]]}')


    async def main_loop(self):
        logger.info('запуск основного цикла планировщика')
        await self.load_schedule_tasks()
        analyzer_task = asyncio.create_task(self._process_analyzer_results())
        file_change_task = asyncio.create_task(self._process_file_changes())
        try:
            while self.running:
                # Busy waiting
                await asyncio.sleep(1)
        except asyncio.CancelledError:
            logger.info('Основной цикл планировщика остановлен')
        finally:
            analyzer_task.cancel()
            file_change_task.cancel()
            for task in self.tasks:
                task.cancel()

    def stop(self):
        self.running = False
        logger.info('Планировщик остановлен')
        



def stop(self):
    pass
